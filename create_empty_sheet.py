from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as oImage
import os
import sys
from PIL import Image
import glob
import pandas as pd
import pprint

width_range = 'BCDEFGHIJKLMNOP'
grey_range = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
blue_fill = PatternFill(start_color='0F6A95',
                   end_color='0F6A95',
                   fill_type='solid')
text_font = Font(name='Calibri',
                 size=9,
                 italic=False,)
title_font = Font(name='Calibri',
                  size=20,
                  bold=True)
title_align = Alignment(horizontal='center',
                        vertical='top')
withe_fill = PatternFill(start_color='FFFFFF',
                   end_color='FFFFFF',
                   fill_type='solid')
grey_fill = PatternFill(start_color='B5BBBF',
                   end_color='B5BBBF',
                   fill_type='solid')
country_map = {"us": "美国",
               "de": "德国",
               "jp": "日本",
               "fr": "法国",
               "gb": "英国",
               "kr": "韩国",
               "cn": "中国",
               "tw": "台湾",}
paid_type_map = {"free": "免费",
                 "paid": "付费",
                 "gross": "畅销"}


def fillcolor(range, startnum, colorfill, sheet):
    for i in range:
        i = i + str(startnum)
        sheet[i].fill = colorfill


def create_title(title, sheet, country, gametype):
    sheet.merge_cells('C5:O7')
    title = '《{}》({}{}榜)'.format(title, country, gametype)
    sheet["C5"] = title
    sheet["C5"].alignment = title_align


def ishan(text):
    for char in text:
        if '\u4e00' <= char <= '\u9fff':
            return True
        break


def shrink_text(text):
    textlen = len(text)
    if ishan(text):
        marklen = 50
    else:
        marklen = 120
    d = ''
    tmp_list = []
    if textlen <= marklen:
        tmp_list.append(text)
    else:
        d = text[marklen]
        while d not in "，。,.:、:' '!?，！。":
            marklen -= 1
            d = text[marklen]
            if marklen == 0:
                return text[:marklen]
                break
        tmp_list.append(text[:marklen+1])
        text = text[marklen+1:]
        text = shrink_text(text)
        tmp_list.extend(text)
    return tmp_list


def clean_text(text):
    new_text = text.replace('<div>', '')
    new_text = new_text.replace('<br>', '')
    new_text = new_text.replace('</div>', '')

    return new_text


def write_descri(sheet, imagefolder):
    des = os.path.join(imagefolder, 'description.txt')
    with open(des) as f:
        lines = f.readlines()
        new_lines = [text.strip() for text in lines]
        new_lines_ = [text for text in new_lines if text != '']
        new_lines = []
        for text in new_lines_:
            if text.startswith('='):
                text = clean_text(text)
                text = text.replace('=', '')
                # text_ = clean_text(text_)
                new_lines.append(text)
            else:
                text = clean_text(text)

                new_lines.append(text)

        tmp = []

        for i in new_lines:
            tmp.extend(shrink_text(i))
        end_line_num = 56 + len(tmp)
        for idx, text in enumerate(tmp, start=56):
            cel = "B" + str(idx)
            sheet[cel] = text
    return end_line_num


def add_icon(sheet, imagefolder):
    icon_path = os.path.join(imagefolder, 'icon.png')
    cell = "G2"
    img = oImage(icon_path)
    sheet.add_image(img, cell)


def add_image(imagefolder, end_num, sheet):
    hor_range = "BGL"
    ver_range = "BEHKN"
    cells = []
    end_num = end_num + 3
    pic_ = imagefolder + "/*.jpg"
    pic_path_list = glob.glob(pic_)
    onepic = pic_path_list[0]

    img_ = Image.open(onepic)
    width_ , height_ = img_.size
    if width_ > height_ :
        for i in hor_range:
            i = i + str(end_num)
            cells.append(i)
        if len(pic_path_list) > 3:
            for i in hor_range[:2]:
                i = i + str(end_num + 12)
                cells.append(i)
        else:
            pass
    else:
        for i in ver_range:
            i = i + str(end_num)
            cells.append(i)
        else:
            pass
    result = zip(pic_path_list, cells)
    for imagepath, cell in result:
        img = oImage(imagepath)
        sheet.add_image(img, cell)


def find_id(imagefolder):
    idfile = imagefolder + "/*_.txt"
    idfilename = glob.glob(idfile)
    id_ = os.path.basename(idfilename[0])
    gameid = id_.rsplit('.', 1)[0].strip('_')
    #gameid = int(gameid)
    return gameid


def checkname(subfolder):
    rightname = []
    subfolder1 = subfolder
    for i in ':?*/：':
        for j in subfolder1:
            if i in j:
                subfolder1.remove(j)
                t = j.replace(i, ' ')
                rightname.append(t)
    rightname.extend(subfolder1)
    return rightname


def check_data(gameid):
    try:
        en_country = app_df[app_df['id'] == gameid]['country'].values[0]
        en_type = app_df[app_df['id'] == gameid]['type'].values[0]
        en_rate = app_df[app_df['id'] == gameid]['rate'].values[0]
        country = country_map.get(en_country)
        cn_type = paid_type_map.get(en_type)
        en_rate = '{:0.1f}'.format(en_rate)
    except:
        country = 'error'
        cn_type = 'error'
        en_rate = 'error'
    finally:
        pass
    return country, cn_type, en_rate


def sheet_main(sheetname, imagefolder):
    ws = wb.create_sheet(title=sheetname)
    gameid = find_id(imagefolder)
    print(gameid)
    game_country, gametype, gamerate = check_data(gameid=gameid)
    for i in range(1, 250):
        fillcolor(grey_range, startnum=i, colorfill=grey_fill, sheet=ws)

    for i in range(2, 180):
        fillcolor(width_range, startnum=i, colorfill=withe_fill, sheet=ws)
    create_title(title=sheetname, sheet=ws, country=game_country, gametype=gametype)


    fix_blue_num = [8, 49, 54]
    for i in fix_blue_num:
        fillcolor(width_range, startnum=i, colorfill=blue_fill, sheet=ws)
    add_icon(sheet=ws, imagefolder=imagefolder)
    end_num = write_descri(sheet=ws, imagefolder=imagefolder)
    end_num = end_num + 3
    end_cel = "B" + str(end_num)
    end_cel1 = "B" + str(end_num + 2)

    ws["B9"] = "游戏下载数量/收入  默认为过去30天内下载/收入"
    ws["B10"] = "IOS收入数据"
    ws["B29"] = "安卓收入数据"
    ws["B50"] = "游戏玩法介绍"
    ws["B55"] = "游戏介绍"
    ws[end_cel] = "游戏评分: {}".format(gamerate)
    ws[end_cel1] = "游戏截图"
    ws["B9"].font = text_font
    ws["B10"].font = text_font
    ws["B29"].font = text_font
    ws["B50"].font = text_font
    ws["B55"].font = text_font
    ws[end_cel].font = text_font
    ws[end_cel1].font = text_font
    fillcolor(width_range, startnum=end_num + 1, colorfill=blue_fill, sheet=ws)
    add_image(imagefolder=imagefolder, end_num=end_num, sheet=ws)


try:
    if sys.argv[1]:

        arg = sys.argv[1]
except:
    arg = '0223'


try:
    if sys.argv[2]:
        filename = sys.argv[2] + '.xlsx'
except:
    filename = '444.xlsx'

wb = Workbook()
path = os.path.expanduser('~/Desktop/{}'.format(arg))
outpath = os.path.expanduser('~/Desktop/')
output_path = os.path.join(outpath, filename)
sub_folder = os.listdir(path)
sub_folder_path = []
app_df = pd.read_excel('/Users/yeye/my_file/senser_tower.xlsx')
app_df['id'] = app_df['id'].astype(str)
if '.DS_Store' in sub_folder:
    sub_folder.remove('.DS_Store')
for i in sub_folder:
    i = os.path.join(path, i)
    sub_folder_path.append(i)
right_sub_folder = checkname(sub_folder)
right_sub_folder.sort()
sub_folder_path.sort()
input_ = zip(right_sub_folder, sub_folder_path)
for gamename, imagefolder in input_:
    print(gamename, imagefolder)
    sheet_main(sheetname=gamename, imagefolder=imagefolder)
wb.save(output_path)
