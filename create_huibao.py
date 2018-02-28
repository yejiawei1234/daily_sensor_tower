from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as oImage
import pandas as pd
import os
from openpyxl.styles import Border, Side
import sys
import glob

if len(sys.argv) < 2:
    print('please give a path')
    sys.exit()
elif len(sys.argv) == 2:
    my_path = sys.argv[1]
else:
    print("you are giving too much argv!!!")
    sys.exit()


border = Border(left=Side(border_style='medium',
                          color='000000'),
                right=Side(border_style='medium',
                           color='000000'),
                top=Side(border_style='medium',
                         color='000000'),
                bottom=Side(border_style='medium',
                            color='000000'),
                )


def add_name(sheet, cell, name):
    sheet[cell] = name


def add_rank(sheet, cell, rank):
    sheet[cell] = rank


def add_up(sheet, cell, up):
    sheet[cell] = up


def add_type(sheet, cell, gametype):
    sheet[cell] = gametype


def add_icon(sheet, cell, imagefolder):
    icon_path = os.path.join(imagefolder, 'small_icon.png')
    img = oImage(icon_path)
    sheet.add_image(img, cell)


country_map = {"us": "美国",
               "de": "德国",
               "jp": "日本",
               "fr": "法国",
               "gb": "英国",
               "kr": "韩国",
               "cn": "中国",
               "tw": "台湾",}
paid_type_map = {"free": "免费",
                 "paid": "付费",
                 "gross": "畅销"}

wb = load_workbook('/Users/yeye/my_file/mou.xlsx')
ws = wb.get_sheet_by_name('总汇报告')
df = pd.read_excel('/Users/yeye/my_file/senser_tower.xlsx')


free_rank_map = {'us': '10',
                 'de': '12',
                 'fr': '14',
                 'jp': '16',
                 'cn': '18',
                 'kr': '20',
                 'ru': '22',
                 'gb': '24',
                 'tw': '26'}

paid_rank_map = {'us': '30',
                 'de': '32',
                 'fr': '34',
                 'jp': '36',
                 'cn': '38',
                 'kr': '40',
                 'ru': '42',
                 'gb': '44',
                 'tw': '46'}

gross_rank_map = {'us': '52',
                  'de': '54',
                  'fr': '56',
                  'jp': '58',
                  'cn': '60',
                  'kr': '62',
                  'ru': '64',
                  'gb': '68',
                  'tw': '70'}

def find_id(imagefolder):
    idfile = imagefolder + "/*_.txt"
    idfilename = glob.glob(idfile)
    id_ = os.path.basename(idfilename[0])
    gameid = id_.rsplit('.', 1)[0].strip('_')
    return gameid

def find_small_icon(imagefolder):
    iconfile = imagefolder + "/small_icon.png"
    return iconfile


def find_sub(filefolder):
    sub_dir_list = [x[0] for x in os.walk(filefolder)][1:]
    return sub_dir_list


home = os.path.join('/Users/yeye/Desktop/', my_path)
sub_dir_list = find_sub(home)
# for i in sub_dir_list:
#     game_id = find_id(i)
#     for j in df.index:
#         game_type = df.at[j, 'type']
#         if game_type == 'free':
#             country = df.at[j, 'country']
#             icon_cell = 'D' + free_rank_map.get(country)
#             add_icon(ws, icon_cell, )
#         elif game_type == 'paid':
#             country = df.at[j, 'country']
#             icon_cell = 'D' + paid_rank_map.get(country)
#             add_icon(ws, icon_cell)
#         else:
#             country = df.at[j, 'country']
#             icon_cell = 'D' + paid_rank_map.get(country)
#             add_icon(ws, icon_cell)

df1 = df.copy()

for i in df.index:
    game_type = df.at[i, 'type']
    if game_type == 'free':
        country = df.at[i, 'country']
        name_cell = 'E' + free_rank_map.get(country)
        rank_cell = 'F' + free_rank_map.get(country)
        up_cell = 'G' + free_rank_map.get(country)
        game_type_cell = 'H' + free_rank_map.get(country)
        add_name(ws, name_cell, df.at[i, 'app_name'])
        add_rank(ws, rank_cell, df.at[i, 'now'])
        add_up(ws, up_cell, df.at[i, 'up'])
        add_type(ws, game_type_cell, df.at[i, '类型'])
        df1.at[i, 'cell'] = 'D' + free_rank_map.get(country)
    elif game_type == 'paid':
        country = df.at[i, 'country']
        name_cell = 'E' + paid_rank_map.get(country)
        rank_cell = 'F' + paid_rank_map.get(country)
        up_cell = 'G' + paid_rank_map.get(country)
        game_type_cell = 'H' + paid_rank_map.get(country)
        add_name(ws, name_cell, df.at[i, 'app_name'])
        add_rank(ws, rank_cell, df.at[i, 'now'])
        add_up(ws, up_cell, df.at[i, 'up'])
        add_type(ws, game_type_cell, df.at[i, '类型'])
        df1.at[i, 'cell'] = 'D' + paid_rank_map.get(country)
    else:
        country = df.at[i, 'country']
        name_cell = 'E' + gross_rank_map.get(country)
        rank_cell = 'F' + gross_rank_map.get(country)
        up_cell = 'G' + gross_rank_map.get(country)
        game_type_cell = 'H' + gross_rank_map.get(country)
        add_name(ws, name_cell, df.at[i, 'app_name'])
        add_rank(ws, rank_cell, df.at[i, 'now'])
        add_up(ws, up_cell, df.at[i, 'up'])
        add_type(ws, game_type_cell, df.at[i, '类型'])
        df1.at[i, 'cell'] = 'D' + gross_rank_map.get(country)

df1['id'] = df1['id'].astype(str)

for i in sub_dir_list:
    game_id = find_id(i)
    cell = df1.loc[df1['id'] == str(game_id), 'cell'].values
    print(cell)
    add_icon(ws, cell[0], i)
    # if cell:
    #     add_icon(ws, cell[0], i)


for row in ws.iter_rows(min_row=10, max_row=27, min_col=3, max_col=9):
    for cell in row:
        cell.border = border

for row in ws.iter_rows(min_row=30, max_row=47, min_col=3, max_col=9):
    for cell in row:
        cell.border = border

for row in ws.iter_rows(min_row=52, max_row=69, min_col=3, max_col=9):
    for cell in row:
        cell.border = border

wb.save('/Users/yeye/Desktop/test123.xlsx')

